import openpyxl

dosya = openpyxl.load_workbook('Liste.xlsx')
sayfa = dosya.active

veri_listesi = []
for satir in sayfa.iter_rows(min_row=2, max_row=792, min_col=1, max_col=1):
    satir_verisi = []
    for hucre in satir:
        satir_verisi.append(str(hucre.value))
    veri_listesi.append(satir_verisi)

# Liste eleman sayısını yazdır
liste_boyutu = len(veri_listesi)
print(f"Liste Eleman Sayısı: {liste_boyutu}")

# 'None' ve 'Code' elemanlarını sil
for satir in veri_listesi:
    if satir == ['None']:
        veri_listesi.remove(satir)

for satir in veri_listesi:
    if satir == ['Code']:
        veri_listesi.remove(satir)
buyuk_harfler = [
    ['A'], ['B'], ['C'], ['D'], ['E'], ['F'], ['G'], ['H'], ['I'], ['J'], ['K'],
    ['L'], ['M'], ['N'], ['O'], ['P'], ['Q'], ['R'], ['S'], ['T'], ['U'], ['V'],
    ['W'], ['X'], ['Y'], ['Z']
]
for satir in veri_listesi:
    if satir in buyuk_harfler:
        veri_listesi.remove(satir)

# Silinmiş listenin boyutunu yazdır
liste_boyutu = len(veri_listesi)
print(f"Silinmiş Liste Eleman Sayısı: {liste_boyutu}")


birden_fazla_elemanli_liste = []

def birden_fazla_eleman(eleman):
  """
  Bir elemanın birden fazla alt elemanı olup olmadığını kontrol eder.

  Args:
    eleman: Kontrol edilecek liste öğesi.

  Returns:
    Eleman birden fazla alt elemana sahipse True, aksi takdirde False.
  """
  if isinstance(eleman, list):
    # Eleman bir listeyse, virgül ile ayrılmış değerleri ayır
    return len(eleman[0].split(", ")) > 1
  else:
    return False

def sil_kisa_elemanlar(birden_fazla_elemanli_liste):
  """
  Birden fazla elemanlı listeden eleman değerlerindeki her bir elemanın karakter uzunluğu 3 ve daha az karakter uzunluğuna sahip öğeleri siler.

  Args:
    birden_fazla_elemanli_liste: İşlenecek liste.

  Returns:
    Kısa eleman değerleri silinmiş yeni liste.
  """
  yeni_liste = []
  for eleman in birden_fazla_elemanli_liste:
    # Eleman bir listeyse ve virgül ile ayrılmış değerleri varsa
    if isinstance(eleman, list):
      tutulacak_elemanlar = []
      for deger in eleman[0].split(", "):  # Her değeri ayrı döngüde işle
        if len(deger) > 3:  # Değer uzunluğu 3'ten fazlaysa tut
          tutulacak_elemanlar.append(deger)
      if tutulacak_elemanlar:  # En az bir eleman tutulursa yeni listeye ekle
        yeni_liste.append(tutulacak_elemanlar)
  return yeni_liste

def sil_kisa_elemanlar_veri_listesi(veri_listesi):
  """
  veri_listesi'nden 3 ve daha az karakter uzunluğuna sahip elemanları siler.

  Args:
    veri_listesi: İşlenecek liste.

  Returns:
    Kısa eleman değerleri silinmiş yeni veri_listesi.
  """
  yeni_veri_listesi = []
  for eleman in veri_listesi:
    # Eleman bir listeyse ve virgül ile ayrılmış değerleri varsa
    if isinstance(eleman, list):
      tutulacak_elemanlar = []
      for deger in eleman[0].split(", "):  # Her değeri ayrı döngüde işle
        if len(deger) > 3:  # Değer uzunluğu 3'ten fazlaysa tut
          tutulacak_elemanlar.append(deger)
      if tutulacak_elemanlar:  # En az bir eleman tutulursa yeni listeye ekle
        yeni_veri_listesi.append([tutulacak_elemanlar])
  return yeni_veri_listesi


yeni_liste = sil_kisa_elemanlar(birden_fazla_elemanli_liste)

# Silinmiş Listeyi Yazdırma
print("# Silinmiş Liste:")
print(veri_listesi)

# veri_listesi'nden kısa elemanları sil
silinmis_veri_listesi = sil_kisa_elemanlar_veri_listesi(veri_listesi)
liste_boyut = len(silinmis_veri_listesi)
print(f"Liste Eleman Sayısı: {liste_boyut}")
# Silinmiş veri_listesi'ni yazdırma
print("# Silinmiş veri_listesi:")
print(silinmis_veri_listesi)
veri_listesi = [item for sublist in silinmis_veri_listesi for item in sublist]
print(veri_listesi)
# Yeni listeyi kaydet
yeni_dosya = openpyxl.Workbook()
yeni_sayfa = yeni_dosya.active

# Verileri yeni sayfaya yazdır
for satir_no, satir_verisi in enumerate(veri_listesi, start=1):
    for sutun_no, veri in enumerate(satir_verisi, start=1):
        yeni_sayfa.cell(row=satir_no, column=sutun_no, value=veri)

# Yeni dosyayı kaydet
yeni_dosya.save('YeniListe.xlsx')

print("Yeni Liste 'YeniListe.xlsx' dosyasına kaydedildi.")